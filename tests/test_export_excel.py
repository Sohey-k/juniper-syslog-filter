"""
test_export_excel.py - Excel出力モジュールのテスト
"""

import pytest
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from modules.export_excel import export_to_excel, ExportExcelError


@pytest.fixture
def temp_csv(tmp_path):
    """テスト用CSVファイルを作成"""
    csv_file = tmp_path / "test_data.csv"
    df = pd.DataFrame(
        {
            "Timestamp": ["2025-12-19T10:00:00Z", "2025-12-19T10:01:00Z"],
            "Hostname": ["srx-fw01", "srx-fw02"],
            "AppName": ["RT_IDP", "RT_IDP"],
            "Severity": ["CRITICAL", "CRITICAL"],
            "Message": [
                "RT_IDP_ATTACK_LOG: Attack detected",
                "RT_IDP_ATTACK_LOG: Intrusion attempt",
            ],
        }
    )
    df.to_csv(csv_file, index=False, encoding="utf-8")
    return csv_file


def test_export_excel_basic(temp_csv, tmp_path):
    """基本的なExcel出力が正しく動作する"""
    output_dir = tmp_path / "output"

    # Excel出力実行
    result = export_to_excel(temp_csv, output_dir, verbose=False)

    # 出力ファイルが存在することを確認
    assert result.exists()
    assert result.suffix == ".xlsx"
    assert result.name == "test_data.xlsx"

    # Excelファイルを読み込んで内容確認
    df = pd.read_excel(result, engine="openpyxl")
    assert len(df) == 2
    assert "Timestamp" in df.columns
    assert "Severity" in df.columns
    assert df["Severity"].tolist() == ["CRITICAL", "CRITICAL"]


def test_export_excel_formatting(temp_csv, tmp_path):
    """フォント・スタイル設定が正しく適用される"""
    output_dir = tmp_path / "output"

    # Excel出力実行
    result = export_to_excel(temp_csv, output_dir, verbose=False)

    # openpyxlで直接確認
    wb = load_workbook(result)
    ws = wb.active

    # ヘッダー（1行目）が太字であることを確認
    for cell in ws[1]:
        assert cell.font.bold is True
        assert cell.font.name == "游ゴシック"
        assert cell.font.size == 11

    # データ行（2行目以降）のフォント確認
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value:
                assert cell.font.name == "游ゴシック"
                assert cell.font.size == 11


def test_export_excel_column_width(temp_csv, tmp_path):
    """列幅が自動調整される"""
    output_dir = tmp_path / "output"

    # Excel出力実行
    result = export_to_excel(temp_csv, output_dir, verbose=False)

    # 列幅確認
    wb = load_workbook(result)
    ws = wb.active

    # すべての列に幅が設定されていることを確認
    for column in ws.columns:
        column_letter = column[0].column_letter
        width = ws.column_dimensions[column_letter].width
        # 最小幅10、最大幅50の範囲内であること
        assert 10 <= width <= 50


def test_export_excel_file_not_found(tmp_path):
    """存在しないファイルを指定した場合、エラーが発生する"""
    output_dir = tmp_path / "output"
    non_existent_file = tmp_path / "non_existent.csv"

    with pytest.raises(ExportExcelError, match="入力ファイルが存在しません"):
        export_to_excel(non_existent_file, output_dir, verbose=False)
