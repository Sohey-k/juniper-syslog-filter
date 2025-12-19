"""
Excel出力モジュール

critical_merged.csv を Excel形式で出力
- フォント: 游ゴシック 11pt
- ヘッダー: 太字
- 列幅: 自動調整
"""

from pathlib import Path
from typing import Union
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font


class ExportExcelError(Exception):
    """Excel出力処理でのエラー"""

    pass


def export_to_excel(
    input_file: Union[str, Path],
    output_dir: Union[str, Path],
    font_name: str = "游ゴシック",
    font_size: int = 11,
    min_width: int = 10,
    max_width: int = 50,
    verbose: bool = True,
) -> Path:
    """
    CSVファイルをExcel形式で出力

    Args:
        input_file: 入力CSVファイルのパス
        output_dir: 出力先ディレクトリ
        font_name: フォント名
        font_size: フォントサイズ
        min_width: 列の最小幅
        max_width: 列の最大幅
        verbose: 詳細ログを出力するか

    Returns:
        Path: 出力されたExcelファイルのPath

    Raises:
        ExportExcelError: Excel出力に失敗した場合

    Examples:
        >>> from pathlib import Path
        >>> output = export_to_excel(
        ...     Path("critical_only/critical_merged.csv"),
        ...     Path("final_output"),
        ...     verbose=True
        ... )
        >>> print(output)
        final_output/critical_merged.xlsx
    """
    input_path = Path(input_file)
    output_dir = Path(output_dir)

    # 入力ファイルの存在確認
    if not input_path.exists():
        raise ExportExcelError(f"入力ファイルが存在しません: {input_path}")

    # 出力ディレクトリの作成
    output_dir.mkdir(parents=True, exist_ok=True)

    # 出力ファイルパス（.csv → .xlsx）
    output_filename = input_path.stem + ".xlsx"
    output_path = output_dir / output_filename

    try:
        # pandasでCSVを読み込み
        df = pd.read_csv(input_path, encoding="utf-8", keep_default_na=False)

        if verbose:
            print(f"  📄 入力: {input_path.name} ({len(df)}行)")

        # Excelに出力（一旦基本的な出力）
        df.to_excel(output_path, index=False, engine="openpyxl")

        if verbose:
            print(f"  📊 Excel出力: {output_path.name}")

        # openpyxlでフォーマット設定
        wb = load_workbook(output_path)
        ws = wb.active

        # フォント設定
        font = Font(name=font_name, size=font_size)
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font

        # ヘッダーを太字に
        header_font = Font(name=font_name, size=font_size, bold=True)
        for cell in ws[1]:
            cell.font = header_font

        # 列幅自動調整
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        # セル値の長さを計算（日本語は2文字分として計算）
                        cell_length = len(str(cell.value))
                        # 日本語文字を含む場合は幅を広げる
                        if any(ord(char) > 127 for char in str(cell.value)):
                            cell_length = int(cell_length * 1.5)
                        max_length = max(max_length, cell_length)
                except:
                    pass

            # 最小幅・最大幅に制限
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 保存
        wb.save(output_path)

        if verbose:
            print(f"  ✨ フォーマット適用完了")
            print(f"  💾 出力: {output_path}")

        return output_path

    except Exception as e:
        raise ExportExcelError(f"Excel出力に失敗しました: {e}")
